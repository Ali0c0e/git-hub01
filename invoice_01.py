# invoice.py

import os
import calendar
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
import pathlib

import csv
import openpyxl, pprint
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import pickle

# 日時を取得
now = date.today()
n_year = now.year
n_month = now.month
n_day = now.day

matubi = calendar.monthrange(n_year, n_month)[1]
jigetu = now + relativedelta(months =+ 1, day = 1)
print('月末:', matubi,  '次月:', jigetu)
j_year = jigetu.year
j_month = jigetu.month
j_day = jigetu.day

# 使用するデータのpathを変数へ
str01 = '年xxxxxaData.xlsx'
data_name = str(n_year) + str01
data_add = pathlib.Path(r'C:/Users/xxxx/xxxx/Data')
data_path = data_add.joinpath(data_name)
print('data_name:', data_name)
print('data_path:', data_path)

str01 = '_' + str(n_month)
month_name = 'xxxxxx_data' + str01 + '.csv'
part_name = 'xxxxx_data' + str01 + '.pkl'
edit_name = 'edit_data' + str01 + '.csv'
edit_add = pathlib.Path(r'C:/Users/xxxxx/xxxxx/xxxxx/output')
out_path = edit_add.joinpath('product.csv')
product_path = edit_add.joinpath('product_df.pkl')
month_path = edit_add.joinpath(month_name)
m_data_path = edit_add.joinpath(part_name)
edit_path = edit_add.joinpath(edit_name)
print('out_path:', out_path)
print('product_path:', product_path)
print('month_path:', month_path)
print('m_data_path:', m_data_path)
print('edit_path:', edit_path)

# 商品一覧と請求書を作成する月受注ページの最終行を取得
data_bk = openpyxl.load_workbook(data_path)
data_sh01 = data_bk['product']
data_sh02 = data_bk[str(n_month)]
end01_r = data_sh01.max_row
end02_r = data_sh02.max_row
end01_r += 1
end02_r += 1
print('end01_r:', end01_r, 'end02_r:', end02_r)

# データを編集する為のBKと使用するSHと作成
sub_bk = openpyxl.Workbook()
sub_bk.create_sheet(index=0, title='product')
sub_bk.create_sheet(index=1, title='sub_data')
sub_bk.create_sheet(index=2, title='part_data')
sub_bk.create_sheet(index=3, title='csv')
sub_bk.sheetnames
sub01_sh = sub_bk['product']
sub02_sh = sub_bk['sub_data']
part_sh = sub_bk['part_data']
csv_sh = sub_bk['csv']
sub01_r = 1
sub02_r = 1
part_r = 1
csv_r = 1

# 商品一覧のデータを1行ごと取得し、sub_bkの準備したsub01_shへ記録する
for data_r in range(2, end01_r):
    if data_sh01.cell(data_r, 2).value is None:
        print('NOT_DATA')
    else:
        for data_col in range(1, 4):
            sub01_sh.cell(sub01_r, data_col).value = data_sh01.cell(data_r, data_col).value
            value = sub01_sh.cell(sub01_r, data_col).value
            print(value)
        sub01_r += 1

# 請求書を作成する受注月のシートのデータをsub_bkのsub02_shへ記録
for data_r in range(10, end02_r):
    if data_sh02.cell(data_r, 3).value is None:
        print('NOT_DATA')
    else:
        for data_col in range(2, 7):
            # 受注データの4列目は不要なためsub02_shへ転記する際、sub02_shの列を詰める
            if data_col == 5 :
                col = data_col - 1
                part_sh.cell(part_r, col).value = data_sh02.cell(data_r, data_col).value
            else:
                col = data_col
                part_sh.cell(part_r, col).value = data_sh02.cell(data_r, data_col).value

            value02 = part_sh.cell(part_r, col).value
            print(value02)
        part_r += 1

# sub_bkのsub01_shに記録した商品データをcsvデータとして記録
with open(out_path, "w", encoding="utf_8_sig") as fp:
    writer = csv.writer(fp, lineterminator="\n")
    for row in sub01_sh.rows:
        writer.writerow([col.value for col in row])
# csvデータを一覧にし、その状態を維持して保存
product_df = pd.read_csv(out_path, encoding="utf_8", header=0, index_col=None)
product_df.to_pickle(product_path)
print(product_df)

# sub_bkのsub02_shに記録した請求書発酵該当データをcsvデータとして記録
with open(month_path, "w", encoding="utf_8_sig") as fp:
    writer = csv.writer(fp, lineterminator="\n")
    for row in part_sh.rows:
        writer.writerow([col.value for col in row])

# csvデータを一覧にし、その状態を維持して保存
part_df = pd.read_csv(month_path, encoding="utf_8", names=['No','code','name','order'], header= None, index_col=None)
part_df.to_pickle(m_data_path)
print(part_df)

# 商品一覧と受注データを[No][code]ごとにグループ化し[order]は加算する
sub_df = part_df.groupby(['No','code'])['order'].sum().reset_index()

# 先ほどグループごとに集計した受注データの左に商品一覧データを[code]を基準に結合
sub_df = sub_df.merge(product_df, how = 'left', on = 'code')
print(sub_df)

# [order]×[価格]を計算し[sale]に値を入れる
sub_df['sale'] = sub_df['order'] * sub_df['price']
print(sub_df)

# 結合、売上の計算までしたデータフレームを新しく項目を指定しout_dfにする
out_df = sub_df.loc[:,['No','name','order','price','sale','code']]
print(out_df)
# out_dfをｃｓｖデータに変換
out_df.to_csv(edit_path)

# 請求書を作成するbkをひらく
bill_name = 'Invoice.xlsx'
bill_add = pathlib.Path(r'C:/Users/xxxx/xxxxxp/*****/****')
bill_path = bill_add.joinpath(bill_name)
print('bill_path:', bill_path)

bill_bk = openpyxl.load_workbook(bill_path)
bill_sh = bill_bk.active

# 新規で編集用のedit_bkを作成し、out_dfをcsvへ変換したデータをExcelデータへ変換し転記
edit_bk = openpyxl.Workbook()
edit = pd.read_csv(edit_path, encoding="utf-8")
edit.to_excel('edit_bk.xlsx', encoding="utf-8")
sh_name = edit_bk.sheetnames

# edit_bkを開き、最終行と最終列を取得
bk = openpyxl.load_workbook('edit_bk.xlsx')
name = bk.sheetnames
sh = bk['Sheet1']
r_end = sh.max_row
c_end = sh.max_column
r_end += 1
c_end += 1

# edit_bkのshのセルに一つ一つ順番に転記していく
for i in range(2, r_end):
    for j in range(2, c_end):
        q = sh.cell(i, j).value
        print(q)

# shのB3を取得し変数b_noへ入れる
c_row = 2
c_col = 3
b_no = sh.cell(c_row, c_col).value
print(b_no)

# 請求書のテンプレに合わせて転記開始の行と列を変数に準備
b_row = 30
b_col = 1
print(b_no)

# shの1行目のデータを請求書シートへ転記する
# 請求書shの8行目のみデータの頭に「№」を付けて転記する
code = sh.cell(c_row, 8).value
bill_sh.cell(b_row, 1).value = str(b_no)
bill_sh.cell(b_row, 2).value = sh.cell(c_row, 4).value
bill_sh.cell(b_row, 5).value = sh.cell(c_row, 5).value
bill_sh.cell(b_row, 6).value = sh.cell(c_row, 6).value
bill_sh.cell(b_row, 7).value = sh.cell(c_row, 7).value
bill_sh.cell(b_row, 8).value = "№" + str(code)

# shの各行の3列目を変数「no」へ入れる
c_row += 1
for i in range(c_row, r_end):
    b_row = i + 28
    b_col = c_col - 2
    no = sh.cell(i, 3).value
    # b_noとnoが同じか調べる
    if b_no == no:
        # 同じ場合は請求書欄の1列目に「〃」を転記
        bill_sh.cell(b_row, b_col).value = " 〃 "
        an01 = bill_sh.cell(b_row, b_col).value
        c_col += 1
        b_col += 1
        # 2列目には商品名を転記する
        if b_col == 2:
            name = sh.cell(i, c_col).value
            bill_sh.cell(b_row, b_col).value = '[＊]' + name
            c_col += 1
            b_col = 5
            # 請求書の5列目以降の転記
            for b_col in range(5, 9):
                # 8列目だったら、shのデータの頭に「№」を付けて転記
                if b_col == 8:
                    code01 = sh.cell(i, c_col).value
                    bill_sh.cell(b_row, b_col).value = "№" + str(code01)
                    c_col = 3
                    b_col = 1
                # 8列目以外はそのまま転記する
                else:
                    bill_sh.cell(b_row, b_col).value = sh.cell(i, c_col).value
                    c_col += 1
                    b_col += 1
    # b_noとnoが異なる場合
    else:
        #sh.cell(i, 3).valueの値をb_noへ入れる
        b_no = sh.cell(i, 3).value
        # 請求書シートのb_rowの1列目にb_noを転記
        bill_sh.cell(b_row, b_col).value = str(b_no)
        c_col += 1
        b_col += 1

        # 2列目には商品名を転記する
        if b_col == 2:
            name = sh.cell(i, c_col).value
            bill_sh.cell(b_row, b_col).value = '[＊]' + name
            c_col += 1
            b_col = 5
            # 請求書の5列目以降の転記
            for b_col in range(5, 9):
                # 8列目だったら、shのデータの頭に「№」を付けて転記
                if b_col == 8:
                    code02 = sh.cell(i, c_col).value
                    bill_sh.cell(b_row, b_col).value = "№" + str(code02)
                    c_col = 3
                    b_col = 1
                # 8列目以外はそのまま転記する
                else:
                    bill_sh.cell(b_row, b_col).value = sh.cell(i, c_col).value
                    c_col += 1
                    b_col += 1

# 以下、請求書のテンプレ（レイアウト)の設定
TITLE_CELL_COLOR = "C0C0C0"
HEADLINE_COLOR = "C0C0C0"

# セルの幅の指定
bill_sh.freez_panes = "C2"
col_widths = {"A":9, "B":3, "C":29, "D":7, "E":7, "F":8, "G":10, "H":12}

for col_name in col_widths:
    bill_sh.column_dimensions[col_name].width = col_widths[col_name]

# 1～12行目のセルの高さを指定
for row_h in range(1, 13):
    bill_sh.row_dimensions[row_h].height = 15

# 27～50行目のセルの高さを指定
for row_h02 in range(27, 51):
    bill_sh.row_dimensions[row_h02].height = 21

# その他のセルの高さを個別に指定
bill_sh.row_dimensions[13].height = 8
bill_sh.row_dimensions[14].height = 8
bill_sh.row_dimensions[15].height = 21
bill_sh.row_dimensions[16].height = 21
bill_sh.row_dimensions[17].height = 13.5
bill_sh.row_dimensions[18].height = 15
bill_sh.row_dimensions[19].height = 18
bill_sh.row_dimensions[20].height = 18
bill_sh.row_dimensions[21].height = 6
bill_sh.row_dimensions[22].height = 25.5
bill_sh.row_dimensions[23].height = 15
bill_sh.row_dimensions[24].height = 20
bill_sh.row_dimensions[25].height = 10
bill_sh.row_dimensions[26].height = 15
bill_sh.row_dimensions[27].height = 24
bill_sh.row_dimensions[51].height = 25.5
bill_sh.row_dimensions[52].height = 15


for x in range(29, 49):
    bill_sh.merge_cells(start_row=x, start_column=2, end_row=x, end_column=4)

# 以下、フォントの指定
font_header = Font(name="コーポレート明朝", size=20, bold=True, color="000000")
font_text = Font(name="コーポレート明朝", size=11, color="000000")
font_s = Font(name="コーポレート明朝", size=10, color="000000")

for i in range(1, 51):
    for j in range(1, 9):
        bill_sh.cell(row=i, column=j).font = Font(name="コーポレート明朝", size=11, color="000000")
        if j == 9:
            bill_sh.cell(row=i, column=j).font = Font(bold=True)

# 数列の場合、「,」を入れるフォーマットを指定
for r_num in range(26, 52):
    for c_num in range(1, 9):
        bill_sh.cell(row=r_num, column= c_num).number_format = "#,##0"
bill_sh.cell(2,27).number_format = "#,##0"

# セルの配置位置を指定
for rx in range(1, 14):
    for cx in range(1, 9):
        bill_sh.cell(row=rx, column=cx).alignment = Alignment(horizontal="left", vertical="center")

# セルの塗りつぶしの指定
for rows in bill_sh["A15:H15"]:
    for rows02 in bill_sh["A16:H16"]:
        for cell in rows:
            for cells in rows02:
                cell.fill = PatternFill(patternType="solid", fgColor=TITLE_CELL_COLOR)
                cells.fill = PatternFill(patternType="solid", fgColor=TITLE_CELL_COLOR)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = font_header

# FAXの送付状部分
bill_sh["A2"] = "株式会社 〇〇〇〇"
bill_sh["A3"].alignment = Alignment(horizontal="left", vertical="center", indent=4)
bill_sh["A3"] = "□□□ 様"

bill_sh["B5"] = "いつもお世話になっております。"
bill_sh["B6"] = ("早速ですが、{}月ご注文いただいた分の請求書を送付させて頂きます。".format(n_month))
bill_sh["B7"] = "内容のご確認よろしくお願い致します。"
bill_sh["B8"] = "内容に不備や訂正等がございましたら、お手数ではございますがお知らせください。"
bill_sh["B9"] = "尚、原本は本日、普通郵便にてお送りさせて頂きます。"
bill_sh["G11"].alignment = Alignment(horizontal="right", vertical="center")
bill_sh["G11"] = "〇〇"
bill_sh["G12"].alignment = Alignment(horizontal="left", vertical="center")
bill_sh["G12"] = "送付枚数 1枚(本状含む)"

# 請求書部分
bill_sh["A15"] = "請　求　書"
bill_sh["G17"].alignment = Alignment(horizontal="right", vertical="top")
bill_sh["G17"] = ("{}年{}月{}日".format(n_year, n_month, matubi))

# 送付先と請求書の内容
bill_sh["A19"].font = Font(name="コーポレート明朝", size=16, color="000000", bold=True)
bill_sh["A19"].alignment = Alignment(horizontal="center", vertical="bottom")
bill_sh["A19"] = "株式会社 〇○○〇"
bill_sh["B22"].alignment = Alignment(horizontal="center", vertical="bottom")
bill_sh["B22"] = ("{}年{}月ご注文分請求書".format(n_year, n_month))

bill_sh["A23"].alignment = Alignment(horizontal="center", vertical="bottom")
bill_sh["A23"] = "件名："
bill_sh["A24"].alignment = Alignment(horizontal="center", vertical="center")
bill_sh["A24"] = "下記の通り、ご請求申し上げます。"

# 請求書の送り主
bill_sh["E22"].alignment = Alignment(horizontal="left", vertical="bottom", indent=2)
bill_sh["E22"].font = Font(name="コーポレート明朝", size=12, color="000000")
bill_sh["E22"] = "株式会社 □□□□□"

bill_sh["E23"].alignment = Alignment(horizontal="left", vertical="bottom", indent=2)
bill_sh["E23"].font = Font(name="コーポレート明朝", size=11, color="000000")
bill_sh["E23"] = "〒000-0000"

bill_sh["E24"].alignment = Alignment(horizontal="left", vertical="top", indent=2)
bill_sh["E24"].font = Font(name="コーポレート明朝", size=11, color="000000")
bill_sh["E24"] = "xx県xx市▲区□丁目〇番地△号"

bill_sh["E25"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
bill_sh["E25"].font = Font(name="コーポレート明朝", size=10, color="000000")
bill_sh["E25"] = "TEL:000-000-0000/FAX:000-000-0000"

bill_sh["E27"].alignment = Alignment(horizontal="left", vertical="top", indent=2)
bill_sh["E27"].font = Font(name="コーポレート明朝", size=11, color="000000")
bill_sh["E27"] = "代表取締役 〇〇 〇〇"

bill_sh["A27"].alignment = Alignment(horizontal="center", vertical="center")
bill_sh["A27"].font = Font(name="コーポレート明朝", size=11, color="000000", bold=True)
bill_sh["A27"] = "合計金額"
bill_sh["C27"].font= Font(name="コーポレート明朝", size=18, color="000000", bold="True")
bill_sh["C27"].alignment = Alignment(horizontal="left", vertical="bottom")
bill_sh["C27"] = '=SUM(G30:G48)'
bill_sh["D27"].alignment = Alignment(horizontal="left", vertical="bottom")
bill_sh["D27"].font = Font(name="コーポレート明朝", size=9, color="000000")
bill_sh["D27"] = " （税込）"

# 請求書の明細部分
for row in bill_sh["H30:H48"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="bottom")
        cell.font = font_text

for col in bill_sh["F49:F51"]:
    for cell in col:
        cell.fill = PatternFill(patternType="solid", fgColor=HEADLINE_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = font_text

bill_sh["A49"] = "小 計"
bill_sh["F50"] = "消費税"
bill_sh["F51"] = "合 計"
bill_sh["G51"] = "=SUM(G30:G48)"

bill_sh["B48"].alignment = Alignment(horizontal="left", indent=2)
bill_sh["B48"] = "※[＊]印は軽減税率「8％」対象商品"

bill_sh["A49"].alignment = Alignment(horizontal="left", vertical="center", indent=4, wrapText=True)
bill_sh["A49"] = "【お振込先】\n〇〇銀行 〇〇支店 普通 １２３４５６\n株式会社□□□□□"


for row2 in bill_sh["A29":"H29"]:
    for cell in row2:
        cell.fill = PatternFill(patternType="solid", fgColor=HEADLINE_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = font_text
# 請求書の項目
bill_sh["A29"] = "№"
bill_sh["B29"] = "品 名"
bill_sh["E29"] = "数 量"
bill_sh["F29"] = "単 価"
bill_sh["G29"] = "金 額"
bill_sh["H29"] = "備 考"

# セルの結合
bill_sh.merge_cells('G1:H1')
bill_sh.merge_cells('A2:D2')
bill_sh.merge_cells('A15:H16')
bill_sh.merge_cells('G17:H18')
bill_sh.merge_cells('G19:H19')
bill_sh.merge_cells('A19:D20')
bill_sh.merge_cells('B19:B20')
bill_sh.merge_cells('C19:C20')
bill_sh.merge_cells('D19:D20')
bill_sh.merge_cells('A24:D25')
bill_sh.merge_cells('G49:H49')
bill_sh.merge_cells('G50:H50')
bill_sh.merge_cells('G51:H51')
bill_sh.merge_cells('A49:E51')
bill_sh.merge_cells('E22:H22')
bill_sh.merge_cells('E23:H23')
bill_sh.merge_cells('E24:H24')
bill_sh.merge_cells('E25:H26')
bill_sh.merge_cells('E27:H27')
bill_sh.merge_cells('A27:B27')
bill_sh.merge_cells('B22:D23')

for col in sh["A30":"A48"]:
    for cell in col:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = font_text
for col2 in sh["H30":"H48"]:
    for cell in col2:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = font_text

# セルの罫線の準備
b_hair = Side(style="hair", color="000000")
b_thin = Side(style="thin", color="000000")
b_thick = Side(style="thick", color="000000")
b_double = Side(style="double", color="000000")
b_dotted = Side(style="dotted", color="000000")

# セルに罫線を引く
border = Border(left=b_thin, right=b_thin, top=b_hair, bottom=b_hair)
border01 = Border(left=b_thin, right=b_thin)
border02 = Border(top=b_thick)
border03 = Border(top=b_thin)
border04 = Border(bottom=b_thick)
border05 = Border(bottom=b_double)
border06 = Border(left=b_thick)
border07 = Border(top=b_thick, bottom=b_thick)
border08 = Border(left=b_thick, right=b_thin, top=b_hair, bottom=b_hair)
border09 = Border(left=b_thin, right=b_thin, top=b_thin, bottom=b_hair)
border10 = Border(left=b_thick, right=b_thin, top=b_thick, bottom=b_thick)
border11 = Border(bottom=b_dotted)
border12 = Border(left=b_thick, top=b_thin)

for z in range(1, 9):
    bill_sh.cell(row=13, column=z).border = border11

for i in range(1, 5):
    bill_sh.cell(row=20, column=i).border = border04
    bill_sh.cell(row=23, column=i).border = border05
    bill_sh.cell(row=27, column=i).border = border05

for row_num in range(29, 51):
    for col_num in range(1, 9):
        bill_sh.cell(row=row_num, column=col_num).border = border

for col_num2 in range(1, 9):
    bill_sh.cell(row=28, column=col_num2).border = border04
    bill_sh.cell(row=49, column=col_num2).border = border03
    bill_sh.cell(row=52, column=col_num2).border = border02


# In[37]:


for q in range(50, 52):
    bill_sh.cell(row=q, column=6).border = border01

for j in range(29, 52):
    bill_sh.cell(row=j, column=1).border = border08
    bill_sh.cell(row=j, column=9).border = border06

for p in range(1, 9):
    bill_sh.cell(row=49, column=p).border = border03
    
for x in range(7, 9):
    bill_sh.cell(row=51, column=x).border = border07


# In[38]:


bill_sh.cell(row=51, column=6).border = border10
bill_sh.cell(row=49, column=6).border = border09
bill_sh.cell(row=49, column=1).border = border12


# In[39]:


bill_bk.save(bill_path)

